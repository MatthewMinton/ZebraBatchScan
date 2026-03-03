# Author: Matthew Minton
# Date: March 2, 2026
#
# Purpose: The purpose of this file is the take the information obtained through scanning from the Zebra_Scan_Tool script
# and appending them to an excel file located in a pre-determined location. This file handles duplicate entries as well as
# empties entries so that errors are minimalized.


import os
import json
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ================= CONFIG =================
TEXT_FILE = r"F:\LFT\98_FCGF_Shares\17_Supply_Chain\End_of_Line\psion.txt"

EXCEL_FILE = r"f:\LFT\98_FCGF_Shares\17_Supply_Chain\End_of_Line\Dock_Scan_Log.xlsx"

CHECKPOINT_FILE = r"F:\LFT\98_FCGF_Shares\17_Supply_Chain\End_of_Line\psion_checkpoint.json"

SHEET_NAME = "Log"
HEADERS = ["Date", "Time", "Serial Number"]
DELIM = ","
# =========================================


def load_checkpoint():
    """
    Returns:
        (offset:int, text_mtime:float)
    """
    if not os.path.exists(CHECKPOINT_FILE):
        return 0, 0.0
    try:
        with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return int(data.get("offset", 0)), float(data.get("text_mtime", 0.0))
    except Exception:
        return 0, 0.0


def save_checkpoint(offset: int, text_mtime: float) -> None:
    os.makedirs(os.path.dirname(CHECKPOINT_FILE), exist_ok=True)
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump(
            {
                "offset": offset,
                "text_mtime": text_mtime,
                "updated": datetime.now().strftime("%m/%d/%Y %H:%M:%S"),
            },
            f,
            indent=2,
        )


def ensure_workbook():
    os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)

    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)

    # Safety check: headers only if sheet is empty
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        ws.append(HEADERS)

    return wb, ws


def load_existing_serials(ws) -> set:
    """
    Loads existing serial numbers from the 'Serial Number' column (column 3),
    skipping the header row.
    """
    serials = set()

    # If sheet is empty or only headers, this will do nothing
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
        val = row[0]
        if val is None:
            continue
        s = str(val).strip()
        if s:
            serials.add(s)

    return serials


def parse_line(line: str):
    """
    Expected format:
    MM/DD/YYYY,HH:MM:SS,18-digit-serial
    """
    line = line.strip()
    if not line:
        return None

    parts = [p.strip() for p in line.split(DELIM)]
    if len(parts) != 3:
        return None

    date_str, time_str, serial = parts

    if len(serial) != 18 or not serial.isdigit():
        return None

    return [date_str, time_str, serial]


def main():
    print("=== PSION TXT -> Excel Daily Append ===")
    print(f"Text file:  {TEXT_FILE}")
    print(f"Excel file: {EXCEL_FILE}")
    print("")

    if not os.path.exists(TEXT_FILE):
        print("ERROR: Text file not found.")
        return

    last_offset, last_mtime = load_checkpoint()

    file_size = os.path.getsize(TEXT_FILE)
    file_mtime = os.path.getmtime(TEXT_FILE)

    # If file got smaller, it was truncated/rotated
    if file_size < last_offset:
        print("NOTICE: Text file truncated, resetting checkpoint.")
        last_offset = 0

    # If file changed since last run but we are exactly at EOF,
    # the producer likely rewrote the file from the beginning.
    if file_mtime > last_mtime and file_size == last_offset:
        print("NOTICE: Text file changed but size matches checkpoint (possible rewrite), resetting checkpoint.")
        last_offset = 0

    new_lines = []
    with open(TEXT_FILE, "r", encoding="utf-8") as f:
        f.seek(last_offset)
        for line in f:
            new_lines.append(line)
        new_offset = f.tell()

    if not new_lines:
        print("No new data found.")
        print("SUCCESS: Completed with 0 new rows.")
        save_checkpoint(last_offset, file_mtime)
        return

    wb, ws = ensure_workbook()

    # ------------------------------------------------------
    # DUPLICATE PROTECTION (BASED ON SERIAL NUMBER IN EXCEL)
    # ------------------------------------------------------
    existing_serials = load_existing_serials(ws)

    rows_to_append = []
    invalid = 0
    duplicates = 0

    for line in new_lines:
        parsed = parse_line(line)
        if not parsed:
            invalid += 1
            continue

        serial = parsed[2]

        # Skip if serial already exists in Excel (or already accepted this run)
        if serial in existing_serials:
            duplicates += 1
            continue

        rows_to_append.append(parsed)
        existing_serials.add(serial)  # prevents duplicates within the same run too

    print(f"New lines read:          {len(new_lines)}")
    print(f"Valid rows to append:    {len(rows_to_append)}")
    if duplicates:
        print(f"Duplicate serials skipped: {duplicates}")
    if invalid:
        print(f"Invalid lines skipped:   {invalid}")

    if not rows_to_append:
        save_checkpoint(new_offset, file_mtime)
        print("No new unique rows, checkpoint updated.")
        print("SUCCESS: Completed with 0 appended rows.")
        return

    start_row = ws.max_row + 1
    for row in rows_to_append:
        ws.append(row)

    wb.save(EXCEL_FILE)
    save_checkpoint(new_offset, file_mtime)

    end_row = ws.max_row
    print("")
    print(f"Appended rows {start_row} to {end_row} in sheet '{SHEET_NAME}'.")

    time.sleep(15)

    print("SUCCESS: Daily append completed.")

    time.sleep(15)

if __name__ == "__main__":
    main()